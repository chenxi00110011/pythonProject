# coding=utf-8
import os
import re
from appium.webdriver.common.touch_action import TouchAction
from appium import webdriver
import environment_variable
import image_properties
import ntp_util
from adjacency_list import LinkedGraph
from image_properties import recognize_text
from my_decorator import retry, timer
from data_store import toDictV5
from environment_variable import ruiboshi_excel, cmri_excel_element_dict
import time
import xrs_adb
import openpyxl
from openpyxl.utils import get_column_letter
from xrs_log import print
import locale
import traceback

locale.setlocale(locale.LC_CTYPE, 'chinese')

'''
1、建议改为显式等待和隐式等待
'''

# 初始化参数
desired_caps = {
    'platformName': 'Android',  # 被测手机是安卓
    'platformVersion': '11',  # 手机安卓版本
    'deviceName': 'xxx',  # 设备名，安卓手机可以随意填写s
    'appPackage': 'com.cmri.universalapp',  # 启动APP Package名称
    'appActivity': '.splash.activity.SplashActivity',  # 启动Activity名称
    'unicodeKeyboard': True,  # 使用自带输入法，输入中文时填True
    'resetKeyboard': True,  # 执行完程序恢复原来输入法
    'noReset': True,  # 不要重置App，如果为False的话，执行完脚本后，app的数据会清空，比如你原本登录了，执行完脚本后就退出登录了
    'newCommandTimeout': 6000,
    'automationName': 'UiAutomator2'
}
WAITTIME = 2


def get_element_data(element, excel_file_path, vertex_name, sheetName='page_element'):
    """
       获取App页面元素数据，并将其存储在Excel文件中
       :param sheetName:
       :param vertex_name:
       :param element:
       :param driver: WebDriver对象
       :param locator: 元素定位表达式，例如(By.ID, 'username')
       :param excel_file_path: Excel文件路径
    """
    try:
        # 打开Excel工作簿
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook[sheetName]
        # 检索最小的空行号
        min_empty_row = 2
        while worksheet[f'A{min_empty_row}'].value:
            min_empty_row += 1
        element_dict = {
            '顶点': vertex_name,
            '邻近点': None,
            'resource_id': element.get_attribute('resource-id'),
            'bounds': element.get_attribute('bounds'),
            'text': element.text,
            'type': None,
            'default_val': None,
            'weight': 1,
            'wait': None
        }
        for idx, key in enumerate(element_dict, start=1):
            col_letter = get_column_letter(idx)
            # print(col_letter,min_empty_row)
            worksheet[f'{col_letter}{min_empty_row}'] = element_dict[key]
        # 保存工作簿
        workbook.save(excel_file_path)

    except Exception as e:
        traceback.print_exc()
        print(e)


class MobieProject:

    def __init__(self, app_name):
        '''
        初始化方法，启动和家亲app，进入设备列表页面
        '''
        if xrs_adb.check_device_connection():  # 检查安卓手机是否已连接
            print('Android device is connected.')
        else:
            raise Exception('No Android device found, or the device is not properly connected.')
        if xrs_adb.check_appium_server(4723):  # 检查appuim服务器端口
            print('Appium server is running.')
        else:
            raise Exception('Appium server is not running.')
        xrs_adb.wakeUpScreen()  # 点亮屏幕
        os.system(xrs_adb.command_dict['滑屏解锁'])
        self.app_name = app_name
        self.app = App()
        self.desired_caps = self.__returnAppPackName(self.app_name, desired_caps)
        self.driver = webdriver.Remote('http://localhost:4723/wd/hub', self.desired_caps)  # 启动appuim客户端
        time.sleep(10)  # 等待app启动
        self.mobile_phone = None
        if app_name == '睿博士':
            self.LinkedGraph = LinkedGraph()  # 创建邻接表
            self.allElemDict = toDictV5(ruiboshi_excel, sheetName='page_element')  # 缓存pwd的页面元素
        elif app_name == '和家亲':
            self.LinkedGraph = LinkedGraph(sheet='universalapp')  # 创建邻接表
            self.allElemDict = toDictV5(ruiboshi_excel, sheetName='universalapp')  # 缓存pwd的页面元素
        # 初始化隐形等待时间
        self.driver.implicitly_wait(10)

    def __returnAppPackName(self, app, desired_caps):
        # 为__init__方法提供self.desired_caps属性
        desired_caps['platformVersion'] = xrs_adb.popen('获取安卓版本')  # 更新手机安卓版本
        desired_caps.update(self.app.app_pack_d[app])  # 加载手机app包名和active
        return desired_caps

    @retry(retries=2, delay=8)  # 当前执行返回未找到元素时，等待一段时间后再次执行
    def clickControl(self, control, mode, wait=None):
        # 按钮控件
        if control is None:
            raise Exception("输入None")
        if mode == 'text':
            self.driver.find_element_by_android_uiautomator(
                f'new UiSelector().text("{control}")'
            ).click()
        elif mode == 'resource_id':
            self.driver.find_element_by_id(control).click()
        elif mode == 'bounds':
            self.driver.tap(control, 300)
        if wait:
            print(f"等待{wait}秒")
            time.sleep(wait)

    @retry(retries=3, delay=8)  # 当前执行返回未找到元素时，等待一段时间后再次执行
    @timer
    def clickControlV1(self, element, content=None):
        '''
        实现页面跳转，为goto方法提供基础方法
        :param element: 实现页面跳转的元素字典
        :param content: 输入框类型，输入的内容
        :return: 无
        '''
        print(f"即将跳转至：\t{element}")  # 打印当前正在跳转的页面元素
        if element['type'] == 'button':  # 按钮类型控件
            for attribute in ['text', 'resource_id', 'bounds']:  # 遍历元素的属性值，判断是否为None，然后执行
                if element[attribute] is None:
                    continue
                elif attribute == 'resource_id':
                    self.driver.find_element_by_id(element[attribute]).click()
                elif attribute == 'text':
                    self.driver.find_element_by_android_uiautomator(
                        f'new UiSelector().text("{element[attribute]}")').click()
                elif attribute == 'bounds':
                    self.driver.tap(self.boundsToCoordinates(element[attribute]), 300)
                break
        elif element['type'] == 'input_box':  # 输入框类型控件
            # # 输入框类型，需要强制等待2s
            # print(f'执行输入')
            # time.sleep(2)
            if content is None:  # 判断content是否为空，否则使用excel中默认值
                content = element['default_val']
            for attribute in ['resource_id', 'text']:
                if element[attribute] is None:  # 遍历元素的属性值，判断是否为None，然后执行
                    continue
                elif attribute == 'resource_id':
                    self.driver.find_element_by_id(element[attribute]).send_keys(content)
                elif attribute == 'text':
                    self.driver.find_element_by_android_uiautomator(
                        f'new UiSelector().text("{element[attribute]}")'
                    ).send_keys(content)
                break
        if element['wait']:
            print(f"等待{element['wait']}秒")
            time.sleep(element['wait'])

    def boundsToCoordinates(self, bounds):
        # 将坐标字符串，转为坐标点。例如[945,123][1035,213]，变为[(x,y)]格式
        physical_size = xrs_adb.get_screen_resolution()
        if physical_size == (1080, 2340):
            deviation_value_x, deviation_value_y = 1, 1
        elif physical_size == (720, 1600):
            deviation_value_x, deviation_value_y = 1.05, 0.938
        result = bounds.split('[')
        result = ','.join(result)
        result = result.split(']')
        result = ''.join(result)
        result = result.split(',')
        x = int((int(result[1]) + int(result[3])) * (physical_size[0] * deviation_value_x / 1080) / 2)
        y = int((int(result[2]) + int(result[4])) * (physical_size[1] * deviation_value_y / 2340) / 2)
        return [(x, y)]

    @retry(retries=2, delay=10)  # 当前执行返回未找到元素时，等待一段时间后再次执行
    def goto(self, *args) -> object:
        # 实现页面跳转，依赖于LinkedGraph类的寻路方法get_road_sign，依赖clickControlV1方法
        # print('开始执行goto')
        currentPageName = self.pwd()  # 执行goto前，先找到当前页面名称
        initial, destination = currentPageName, args[0]
        if initial == destination:  # 判断起始点与终点是否相同，相同则返回
            return
        emel_list = self.LinkedGraph.get_road_sign(initial, destination)  # 寻路
        i = 1
        for emel in emel_list:
            # print(args)
            if emel['type'] == 'input_box' and len(args) >= 2:  # 判断控件类型和形参个数
                content = args[i]
                i += 1
            else:
                content = None
            self.clickControlV1(emel, content)
        # print('结束执行goto')

    @retry(retries=2, delay=8)  # 当前执行返回未找到元素时，等待一段时间后再次执行
    def enterTo(self, control, content, mode):
        # 输入框控件
        if control is None:
            raise Exception("输入None")
        if mode == 'text':
            self.driver.find_element_by_android_uiautomator(
                f'new UiSelector().text("{control}")'
            ).send_keys(content)
        elif mode == 'resource_id':
            self.driver.find_element_by_id(control).send_keys(content)
        # elif mode == 'bounds':
        #     self.driver.tap(control, 300)

    def is_element_exist(self, element, times=3, wait=0, page_source=None):
        """验证页面是否存在某个元素，用于判断页面是否跳转成功"""
        # 判断输入是否为int类型，是则转变为str类型
        if isinstance(element, int):
            element = str(element)
        count = 0
        while count < times:
            if page_source is None:
                souce = self.driver.page_source
            else:
                souce = page_source
            # print(element)
            if re.search(element, souce):  # 通过正则表示判断
                return True
            # if element in souce:
            #     return True
            else:
                count += 1
                time.sleep(wait)
        return False

    @timer
    def pwd(self):
        """找到当前所处页面"""
        page_source = self.driver.page_source  # 将is_element_exist方法需要用的页面元素提前缓存，优化运行效率
        if not xrs_adb.is_foreground(self.desired_caps['appPackage']):  # 判断app是否前台运行
            raise Exception('未检测到指定app')
        allElemDict = self.allElemDict  # 将pwd方法需要用的页面元素提前缓存，优化运行效率
        degreeOfRealism = dict()
        for pageName, elemList in allElemDict.items():
            degreeOfRealism[pageName] = 0
            for elem in elemList:
                # print(elem)   # 出现有的元素通过不了正则表达式
                for attribute in ['text', 'resource_id']:
                    if elem[attribute] is None:
                        continue
                    # 增加page_source形参是为了优化方法用时太长问题
                    elif self.is_element_exist(elem[attribute], times=1, page_source=page_source):
                        degreeOfRealism[pageName] += elem['weight']  # 加权
        for pageName, score in degreeOfRealism.items():  # 通过分数判断当前页面
            # print(pageName,':', score)  # 打印页面评分
            if score == max(degreeOfRealism.values()):
                print(f'当前页面为<{pageName}>')
                return pageName

    def scroll_to_element(self, element):
        """滑动屏幕找到元素element"""
        size = self.driver.get_window_size()
        # 当我第一次进入页面的时候：
        found = False
        count = 0
        old = None
        new = self.driver.page_source
        while not found and count != 1:
            if old == new:
                count += 1
            else:
                # 找元素
                if self.driver.page_source.find(element) != -1:
                    # print('找到了对应的内容')
                    found = True
                else:
                    # 找不到元素的时候，滑动，此时页面更新
                    self.driver.swipe(size['width'] * 0.5, size['height'] * 0.8, size['width'] * 0.5, size['height'] * 0.2, 200)
                    time.sleep(2)
                    # 更新old 的值。用new 的值更新old 的值
                    old = new
                    # 更新new 的值为滑动后的page_source
                    new = self.driver.page_source
        return found

    def long_press_element_by_uiautomator(self, selector, duration=2000):
        """
        在 Appium 中封装长按元素方法。

        Args:
            driver: WebDriver 实例
            selector: 需要长按的 WebElement 对象的 UiSelector，可以是字符串或者 UiSelector 对象
            duration: 长按持续时间，默认为 2000 毫秒

        Returns:
            None
        """
        if isinstance(selector, str):
            selector = f'new UiSelector().text("{selector}")'
        element = self.driver.find_element_by_android_uiautomator(selector)
        action = TouchAction(self.driver)
        action.long_press(element).wait(duration).release().perform()

    def find_nearest_element(self, attribute: dict, text_value):
        """通过resource_id定位多个元素，找到距离文本为text_value的元素最近的一个"""
        if 'resource_id' in attribute.keys():
            # 找到所有具有给定 resource-id 的元素
            elements = self.driver.find_elements_by_id(attribute['resource_id'])
        elif 'text' in attribute.keys():
            attribute_text = attribute['text']
            elements = self.driver.find_elements_by_android_uiautomator(
                f'new UiSelector().text("{attribute_text}")')

        # 获取参考元素的坐标
        if type(text_value) == type('string'):
            reference_element = self.driver.find_element_by_android_uiautomator(
                f'new UiSelector().text("{text_value}")')
        else:
            reference_element = text_value
        ref_location = reference_element.location

        # 计算其他元素的中心坐标并选择最接近参考元素的元素
        nearest_element, nearest_distance = None, float('inf')
        for element in elements:
            element_location = element.location
            element_size = element.size
            element_center = (element_location['x'] + element_size['width'] / 2,
                              element_location['y'] + element_size['height'] / 2)
            distance = ((element_center[0] - ref_location['x']) ** 2 +
                        (element_center[1] - ref_location['y']) ** 2) ** 0.5
            # print(element,element_center,distance,ref_location)
            if distance < nearest_distance:
                nearest_element, nearest_distance = element, distance

        return nearest_element

    def is_element_checkable(self, element):
        """判断元素是否勾选，true表示已勾选"""
        checkable_attr = element.get_attribute('checked')
        return checkable_attr == 'true'

    def pinch_zoom(self, start, end, scale_factor=1.25, duration=500):
        action = TouchAction(self.driver)

        # 计算中心点
        start_x, start_y = start
        end_x, end_y = end
        mid_x = (start_x + end_x) // 2
        mid_y = (start_y + end_y) // 2

        # 第一根手指按下屏幕
        action.press(x=start_x, y=start_y)

        # 第二根手指按下屏幕并滑动到位置
        x1 = int(end_x + (end_x - mid_x) * (scale_factor - 1) / 2)
        y1 = int(end_y + (end_y - mid_y) * (scale_factor - 1) / 2)
        action.move_to(x=x1, y=y1).wait(duration)

        # 两根手指同时离开屏幕
        action.move_to(x=mid_x, y=mid_y).release()
        action.perform()

    def check_sdcard_recording_breakpoint(self, time_slot):
        """睿博士卡录像断点判断"""
        # 判断是否在回放页
        if self.pwd() != '回放页':
            raise Exception('当前页面不是回放页')
        os.system(xrs_adb.command_dict['手机截屏'])
        os.system(xrs_adb.command_dict['下载手机截屏'])

        # 截图，将回放进度条图片转变为时间字符串
        text_coords = recognize_text('C:\\Users\\Administrator\\Desktop\\video\\Screenshots\\screenshot.png',
                                     (0.44, 0.52, 0, 1))
        print(text_coords.keys())
        time_list = text_coords.keys()

        # 过滤非法时间，并返回时间区间
        time_list = [time for time in time_list if re.match(r'^([01][0-9]|2[0-3]):[0-5][0-9]$', time)]
        time_range = (time_list[0], time_list[-1])

        # 获取屏幕分辨率，找到进度条的Y轴值
        screen_resolution = xrs_adb.get_screen_resolution()
        ruleview_y = screen_resolution[1] * 0.48

        # 判断需要检查的时间区间是否在页面中，不在则对应左移或者右移
        if ntp_util.compare_time(time_slot[0], time_range[0]):
            touch = TouchAction(self.driver)
            touch.press(x=100, y=ruleview_y).wait(1000).move_to(x=540, y=ruleview_y).release().perform()
            return self.check_sdcard_recording_breakpoint(time_slot)
        elif not ntp_util.compare_time(time_slot[0], time_range[1]):
            touch = TouchAction(self.driver)
            touch.press(x=900, y=ruleview_y).move_to(x=540, y=ruleview_y).release().perform()
            return self.check_sdcard_recording_breakpoint(time_slot)

        # 将时间换算为坐标X值，然后获取该区间内的所有像素的RGB值
        print('找到对应的区间')
        first_coordinate = (text_coords[time_range[0]][0], text_coords[time_range[1]][0])
        print(time_range, time_slot, first_coordinate)
        second_coordinate = ntp_util.get_coordinates(time_range, time_slot, first_coordinate)
        pixel_info = dict()
        for x in range(second_coordinate[0], second_coordinate[1], 1):
            pixel_info[(x, 80)] = image_properties.get_pixel_color('cropped.jpg', (x, 80))
        # print(pixel_info)

        # 将区间拖到中心，播放该区间的卡录像
        touch = TouchAction(self.driver)
        touch.press(x=second_coordinate[0], y=ruleview_y).wait(1000).move_to(x=540, y=ruleview_y).release().perform()

        # 遍历该RGB值的字典，判断是否有断点
        result = []
        for pixel, rgb in pixel_info.items():
            if sum(rgb) < 10:
                result.append(pixel)
        if result:
            return False
        else:
            return True


class App:
    # app包名
    app_pack_d = {
        '和家亲': {'appPackage': 'com.cmri.universalapp', 'appActivity': '.splash.activity.SplashActivity'},
        '睿博士': {'appPackage': 'com.zwcode.p6slite', 'appActivity': '.activity.SplashActivity'}
    }


'''
1、完善邻接表
2、获取当前位置的方法
'''
if __name__ == "__main__":
    # # print(a.appIsRunning('com.zwcode.p6slite'))
    # print(a.get_deviceid() == [])
    # # # print(a.appIsRunning("adb shell 'dumpsys window | grep mCurrentFocus'"))

    # 打印当前位置
    # project = MobieProject('和家亲')
    # while True:
    #     project.pwd()
    #     time.sleep(5)
    # ruiboshi = MobieProject('睿博士')
    # ruiboshi.goto('报警管理页')
    # element = ruiboshi.find_nearest_element('com.zwcode.p6slite:id/param_switch', '移动侦测报警')
    # print(ruiboshi.is_element_checkable(element))
    #############################################################################
    # ruiboshi = MobieProject('睿博士')
    # ruiboshi.goto('回放页')
    # time.sleep(10)
    # ruiboshi.pinch_zoom((200,1065),(800,1065))
    # """检查SD卡回放条"""
    # ruiboshi = MobieProject('睿博士')
    # ruiboshi.goto('回放页')
    # time.sleep(10)
    # print(ruiboshi.check_sdcard_recording_breakpoint(('11:00', '12:00')))
    """"""
    project = MobieProject('睿博士')
    project.goto('手动添加页')
    project.goto('首页', 'IOTDBB-065896-UXLYD', 'test_dev')