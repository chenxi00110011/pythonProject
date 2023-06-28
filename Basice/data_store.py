import json
import os
import time
import openpyxl
from openpyxl import load_workbook
from environment_variable import ruiboshi_excel
from my_decorator import timer

dataFile = "./date_dict.txt"
data_d = {
    '设备信息': {
        '设备型号': 'xxx',
        '版本号': 'xxx',
        '串口号': 'COM35'
    },
    '运行环境': {
        '测试场景': '',
        '视频存储方式': '卡存储/云存储'
    },
    '异常': {
        '重启': '1',
        '升级': ''
    },
    '测试结果': {}
}


def readToFile():
    # 读数据
    if os.path.exists(dataFile):
        with open(dataFile, 'r') as file_to_read:
            loaded_simple_dict = json.load(file_to_read)
            return loaded_simple_dict
    else:
        return {}


def writeToFile(key, value):
    '''
    对 "./date_dict.txt"文件追加数据
    :param key: 数据的键
    :param value: 数据的值
    :return: 无
    '''
    data_dict = readToFile()
    if key in data_dict.keys() and data_dict[key]:
        data_dict.update({key: data_dict[key] + [value]})
    else:
        data_dict[key] = [value]
    with open(dataFile, 'w') as file_to_write:
        # 进行json序列化,然后写入simple_dict.txt文件中
        json.dump(data_dict, file_to_write)


def tabulation_write(excel_url, coordinates, data):
    '''
    excel文件写入数据
    :param excel_url: excel文件的路径
    :param coordinates: 单元格位置，例如"A1"
    :param data: 写入数据，如果是字符串则直接写入，如果是字典或列表则转成字符串再写入
    :return: 无
    '''
    data = dictToStr(data)
    print(data)
    book = load_workbook(excel_url)
    sheet = book.active
    sheet[coordinates] = data
    # sheet.cell(row=2, column=2).value = 111
    book.save(excel_url)


def dictToStr(_dict):
    '''
    将数据转成字符串
    :param _dict:可以传入字典，列表，字符串
    :return:字符串类型数据
    '''
    if type(_dict) == dict:
        value = ''
        # print("判断成功")
        for k, v in _dict.items():
            value += f'{k}:{dictToStr(v)}\n'
        # print(value)
    elif type(_dict) == list:
        value = '\n'.join(_dict)
    else:
        # print(type(_dict) ,"判断失败")
        value = _dict
    return value


def write_dict(excel_url, data_d):
    '''
    此方法用于，规定写入的位置，目前暂无特别好的方案
    :param excel_url: excel文件路径
    :param data_d: 写入数据，该数据为字典类型
    :return: 无
    '''
    for key in data_d:
        if key == '设备信息':
            tabulation_write(excel_url, 'A3', data_d[key])
        elif key == '运行环境':
            tabulation_write(excel_url, 'B3', data_d[key])
        elif key == '异常':
            tabulation_write(excel_url, 'C3', data_d[key])
        elif key == '测试结果':
            tabulation_write(excel_url, 'D3', data_d[key])


def toDict(excel_url):
    '''
    将excel文件转成字典
    :param excel_url: excel文件路径
    :return: 字典
    '''
    return_dict = dict()
    book = load_workbook(excel_url)
    sheet = book.active
    if sheet['A1'] is None:
        raise Exception("文件无数据，请检查文件路径是否正确或检查文件数据是否正确")
    i, j = 2, 66  # 2表示从第二行开始，66代表B的acsii码
    while sheet[f'{chr(j)}1'].value:
        i = 2
        return_dict[sheet[f'{chr(j)}1'].value] = dict()
        print(sheet[f'{chr(j)}1'].value)
        while sheet[f'A{i}'].value:
            print(sheet[f'A{i}'].value)
            return_dict[sheet[f'{chr(j)}1'].value][sheet[f'A{i}'].value] = sheet[f'{chr(j)}{i}'].value
            print(f"第{i}行，对应值为{sheet[f'{chr(j)}{i}'].value}")
            time.sleep(1)
            i += 1
        j += 1
        print(return_dict)
    return return_dict


def toDictV1(excel_url):
    '''
    将excel文件转成字典
    :param excel_url: excel文件路径
    :return: 字典
    '''
    return_dict = dict()
    book = load_workbook(excel_url)
    sheet = book.active
    if sheet['A1'] is None:
        raise Exception("文件无数据，请检查文件路径是否正确或检查文件数据是否正确")
    i, j = 1, 66  # 2表示从第二行开始，66代表B的acsii码
    while sheet[f'A{i}'].value:
        # print(sheet[f'A{i}'].value)
        return_dict[sheet[f'A{i}'].value] = sheet[f'{chr(j)}{i}'].value
        # print(f"第{i}行，对应值为{sheet[f'{chr(j)}{i}'].value}")
        # time.sleep(1)
        i += 1
        # print(return_dict)
    return return_dict


def toDictV2(excel_url, pageName, sheet):
    '''
    将dict_ruiboshi.xlsx中的页面元素转变成邻接表（字典）
    {页面名称:{resource_id,bounds,text,type,default_val}}
    :param sheet: excel的sheet页名称
    :param pageName: 页面名称，找出符合条件
    :param excel_url: excel文件路径
    :return: 字典
    '''
    return_dict = dict()
    book = load_workbook(excel_url)
    # sheet = book.active
    sheet = book[sheet]
    if sheet['A1'] is None:
        raise Exception("文件无数据，请检查文件路径是否正确或检查文件数据是否正确")
    i, j = 2, 67  # i表示行号，j表示列号（ascii码）
    while sheet[f'A{i}'].value:
        j = 66
        if sheet[f'A{i}'].value == pageName:
            if not sheet[f'B{i}'].value:
                i += 1
                continue
            if sheet[f'B{i}'].value not in return_dict:  # 判断页面跳转元素是否已添加，未添加则初始化为[]
                return_dict[sheet[f'B{i}'].value] = []
            return_dict[sheet[f'B{i}'].value].append(dict())
            while sheet[f'{chr(j)}1'].value:
                return_dict[sheet[f'B{i}'].value][-1][sheet[f'{chr(j)}1'].value] = sheet[f'{chr(j)}{i}'].value
                j += 1
        i += 1
    return return_dict


def toDictV3(excel_url):
    '''
    将dict_ruiboshi.xlsx中的页面名称转变成列表
    :param excel_url: excel文件路径
    :return: 字典
    '''
    return_list = list()
    book = load_workbook(excel_url)
    sheet = book['page_element']
    if sheet['A1'] is None:
        raise Exception("文件无数据，请检查文件路径是否正确或检查文件数据是否正确")
    i = 2  # i表示行号
    while sheet[f'A{i}'].value:
        if sheet[f'A{i}'].value not in return_list:
            return_list.append(sheet[f'A{i}'].value)
        if sheet[f'B{i}'].value not in return_list and sheet[f'B{i}'].value:
            return_list.append(sheet[f'B{i}'].value)
        i += 1
    return return_list


def toDictV4(excel_url):
    '''
    将dict_ruiboshi.xlsx中的邻接表转变成列表
    :param excel_url: excel文件路径
    :return: list
    '''
    return_list = list()
    book = load_workbook(excel_url)
    # sheet = book.active
    sheet = book['page_element']
    if sheet['A1'] is None:
        raise Exception("文件无数据，请检查文件路径是否正确或检查文件数据是否正确")
    i = 2  # i表示行号，66代表B的acsii码
    while sheet[f'A{i}'].value:
        if sheet[f'B{i}'].value:
            return_list.append([sheet[f'A{i}'].value])
            return_list[-1].append(sheet[f'B{i}'].value)
        i += 1

    # 去重
    new_li = []
    for i in return_list:
        if i not in new_li:
            new_li.append(i)
    return new_li


def toDictV5(excel_url, sheetName='page_element'):
    '''
    将dict_ruiboshi.xlsx中的页面元素转变成邻接表（字典）
    {页面名称:{resource_id,bounds,text,type,default_val}}
    :param sheetName:
    :param excel_url: excel文件路径
    :return: 字典
    '''
    return_dict = dict()
    book = load_workbook(excel_url)
    # sheet = book.active
    sheet = book[sheetName]
    if sheet['A1'] is None:
        raise Exception("文件无数据，请检查文件路径是否正确或检查文件数据是否正确")
    i, j = 2, 67  # i表示行号，j表示列号（ascii码）
    while sheet[f'A{i}'].value:
        j = 66
        if sheet[f'A{i}'].value not in return_dict:  # 判断页面跳转元素是否已添加，未添加则初始化为[]
            return_dict[sheet[f'A{i}'].value] = []
        return_dict[sheet[f'A{i}'].value].append(dict())
        while sheet[f'{chr(j)}1'].value:
            return_dict[sheet[f'A{i}'].value][-1][sheet[f'{chr(j)}1'].value] = sheet[f'{chr(j)}{i}'].value
            j += 1
        i += 1
    return return_dict


def read_excel_to_dict(excel_path, worksheet_name, did, did_col=0):
    # 加载 Excel 文件
    wb = openpyxl.load_workbook(excel_path)

    # 获取指定工作表
    sheet = wb[worksheet_name]

    # 获取表头行，并创建字典键
    headers = [cell.value for cell in sheet[1]]
    result = {header: [] for header in headers}

    # 读取数据，填充字典值
    for row in sheet.iter_rows(min_row=2, values_only=True):

        if row[did_col] == did:
            for header, value in zip(headers, row):
                result[header].append(value)

    # 返回字典结果
    return result

if __name__ == '__main__':
    # print(read_excel_to_dict(ruiboshi_excel, '设备详情', 'IOTDBB-065896-UXLYD', 3))
    # new_dict = {'COM5':'456'}
    # writeToFile(new_dict)
    print(toDictV2(ruiboshi_excel, pageName='首页',sheet='page_element'))
    # tabulation_write(url,'C3',d)
    # print(dictToStr(data_d))